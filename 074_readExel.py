# как считывать данные с EXEL таблицы

import openpyxl

book = openpyxl.open("074_data.xlsx", read_only=True)  # открываем таблицу для чтения
sheet = book.active  # указали что читаем первый лист

print(sheet["B2"].value)  # выибираем ячейку
print(sheet[1][2].value)  # 1-колонка , 2-ряд(но начинается с 0) = 3-ряд

print(sheet[1][0].value)  # 1-колонка , 1-ряд(но начинается с 0) = 2-ряд

print('\n\n Получаем все столбцы от 1 6 строки:\n')

for row in range(1, 7):
    print(sheet[row])

print('\n\n Получаем значения всех столбов и строчек таблицы:\n')

for row in range(1, sheet.max_row + 1):
    author = sheet[row][0].value  # value чтобы получить значение ячейки
    name = sheet[row][1].value
    year = sheet[row][2].value
    rate = sheet[row][3].value
    print(row, author, name, year, rate)

print("\n\n Работаем с диапозонами b1-c7:\n")

сells = sheet['B2':'C7']
for cell in сells:
    print(cell)
print("Получили кортеж \n")

print("\nРаспаковываем кортеж:\n")

сells = sheet['B2':'C7']
for name, yearss in сells:
    print(name.value, yearss.value)

print("\n\n используем метод iter_rows для перебора строк:\n")

for row in sheet.iter_rows(min_row=2, max_row=5, min_col=1,max_col=3):  # если значения в iter_rows убрать то мы
    # обойдем все значения как max_row и max_colm
    for cell in row:
        print(cell.value, end=' ')  # end отменяет вывод в столбик и делает его в строчку
    print()  # после выполнения ряда мы переходим на новую строчку

print("\n\n переключаемся между листами в EXEL таблице:\n")

sheets = book.worksheets[2]
print(sheets['A1'])
